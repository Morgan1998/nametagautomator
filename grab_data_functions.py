from openpyxl import load_workbook

def get_final_row(ws):
    i = 1
    while ws.cell(row=i, column=1).value is not None:
        i += 1
    else:
        final_row = i - 1
        return final_row

def get_group_count(ws):
    final_row = get_final_row(ws)
    group_count = []
    for i in range(3, final_row + 2):
        if (len(str(ws.cell(row=i, column=1).value)) > 2 and len(str(ws.cell(row=i + 1, column=1).value)) > 2 and ws.cell(row=i, column=1).value is not None) or (ws.cell(row=i, column=1).value is None and len(str(ws.cell(row=i - 1, column=1).value)) < 3): # Make sure this will always work
            group_count_number = range(len(group_count) + 1)
            group_count.append(f"group {group_count_number}") # make simpler way to do this (such as group_count+=)
    return len(group_count)

def get_end_row(ws, group):
    final_row = get_final_row(ws)
    end_row = None
    group_counter = 0
    for i in range(1, final_row + 2):
        if (len(str(ws.cell(row=i, column=1).value)) < 3 and len(str(ws.cell(row=i + 1, column=1).value)) > 2) or (len(str(ws.cell(row=i, column=1).value)) < 3 and len(str(ws.cell(row=i + 1, column=1).value)) is None):
            group_counter += 1
        if group_counter == group:
            end_row = i
            break
    return end_row

def get_start_row(ws, group):
    final_row = get_final_row(ws)
    if group == 1:
        start_row = 3
        return start_row
    else:
        start_row = None
        group_counter = 0
        for i in range(1, final_row + 2):
            if len(str(ws.cell(row=i, column=1).value)) < 3 and len(str(ws.cell(row=i - 1, column=1).value)) > 2:
                group_counter += 1
            if group_counter == group:
                start_row = i
                break
        return start_row

def get_fn_sc_list(ws, name_col, school_col, start_row, end_row): # fn = first name; sc = school code
    full_names = [c.value for c in (ws[name_col][(start_row - 1):end_row])]
    school_codes = [c.value for c in (ws[school_col][(start_row - 1):end_row])]
    first_names = []
    for full_name in full_names:
        split_names = []
        split_names.append(full_name.split())
        for first_name_last_name in split_names:
            first_names.append(first_name_last_name[0])
    fn_sc_list = []
    for first_name, school_code in zip(first_names, school_codes):
        fn_sc_list.append([first_name, school_code])
    return fn_sc_list

def combine_firstname_with_school_code_by_group(wb_filename: str, ws_name: str, group_count: int, names_col: str, school_code_col: str):
    wb = load_workbook(wb_filename)
    ws = wb[ws_name]
    full_list = []
    for group in range(1, group_count + 1):
        full_list.append(get_fn_sc_list(ws, names_col, school_code_col, get_start_row(ws, group), get_end_row(ws, group)))
    return full_list

