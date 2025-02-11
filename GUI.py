import dearpygui.dearpygui as dpg
from openpyxl import load_workbook
from grab_data_functions import get_group_count, combine_firstname_with_school_code_by_group
from nametag_gen_func import generate_name_tags

dpg.create_context()
dpg.create_viewport(title='Custom Title', width=650, height=1000)

excel_file_path = []
work_sheet_name = []
group_count = []
path_and_sheet_and_count = [excel_file_path, work_sheet_name, group_count]

current_back_side_image = None
current_logo = None

current_group_options_buttons = []
current_group_image_dialogs = []
current_images = []
current_group_windows = []
user_train_options = []
preview_train_options = []
user_name_options = []
preview_name_options = []



# Main functions

def create_name_tags(sender, app_data, user_data):
    if excel_file_path == []:
        pass
    else:
        path = user_data[0][0]
        ws_name = user_data[1][0]
        group_count = user_data[2][0] # group_count is group_count
        full_list = combine_firstname_with_school_code_by_group(path, ws_name, group_count, "D", "I")
        groups_and_images_zipped = []
        for x, y in zip(full_list, current_images):
            groups_and_images_zipped.append(x)
            groups_and_images_zipped.append(y)
        generate_name_tags(group_count, user_train_options, user_name_options, current_back_side_image, current_logo, *groups_and_images_zipped)

def selected_excel_file_callback(sender, app_data):
    global current_group_options_buttons# find out better way to edit global scope lists (Also, test out making them global at top of module)
    global current_group_image_dialogs
    global current_images
    global current_group_windows
    global user_train_options
    global preview_train_options
    global user_name_options
    global preview_name_options

    default_train_text_size = 20
    default_name_text_size = 36

    if excel_file_path == []:
        wb = load_workbook(app_data['file_path_name'])
        ws = wb.worksheets[0]
        excel_file_path.append(app_data['file_path_name'])
        work_sheet_name.append(wb.sheetnames[0]) # figure out why you did this
        group_count.append(get_group_count(ws))
        group_count_range = group_count[0] + 1 # group_count range
        current_images = [i for i in range(1, group_count_range)]

        
        if current_group_options_buttons == [] and current_group_image_dialogs == []:
            current_group_windows = [i for i in range(1, group_count_range)]

            for i in range(1, group_count_range):
                user_train_options.append(['OPTIVagRound-Bold', '#000000', True, default_train_text_size])
            for i in range(1, group_count_range):
                user_name_options.append(['OPTIVagRound-Bold', '#000000', True, default_name_text_size])
            for i in range(1, group_count_range):
                preview_train_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_train_text_size])
            for i in range(1, group_count_range):
                preview_name_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_name_text_size])

            make_dialogs_for_choosing_group_image(group_count_range)
            make_group_options_buttons(group_count_range)
            dpg.add_texture_registry(tag='texture registry', show=False)
            
        else: # Do you even need this else section?
            for button in current_group_options_buttons: # make delete buttons function
                dpg.delete_item(button)
                dpg.delete_item(button + " spacer")
            for index, window in enumerate(current_group_windows):
                dpg.delete_item(window)

            user_train_options.clear()
            for i in range(1, group_count_range):
                user_train_options.append(['OPTIVagRound-Bold', '#000000', True, default_train_text_size])
            user_name_options.clear()
            for i in range(1, group_count_range):
                user_name_options.append(['OPTIVagRound-Bold', '#000000', True, default_name_text_size])
            preview_train_options.clear()
            for i in range(1, group_count_range):
                preview_train_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_train_text_size])
            preview_name_options.clear()
            for i in range(1, group_count_range):
                preview_name_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_name_text_size])

            current_group_windows = [i for i in range(1, group_count_range)]

            for dialog in current_group_image_dialogs:
                dpg.delete_item(dialog)

            dpg.delete_item('create')
            dpg.delete_item('texture registry')
            current_group_image_dialogs.clear()
            current_group_options_buttons.clear()
            make_dialogs_for_choosing_group_image(group_count_range)
            make_group_options_buttons(group_count_range)
            dpg.add_texture_registry(tag='texture registry', show=False)

    else:
        excel_file_path.clear()
        work_sheet_name.clear()
        group_count.clear()
        wb = load_workbook(app_data['file_path_name'])
        ws = wb.worksheets[0]
        excel_file_path.append(app_data['file_path_name'])
        work_sheet_name.append(wb.sheetnames[0]) # Find out if you want to use this
        group_count.append(get_group_count(ws))
        group_count_range = group_count[0] + 1  # group_count range
        current_images = [i for i in range(1, group_count_range)]

        if current_group_options_buttons == [] and current_group_image_dialogs == []: # Is this if statement even needed? Maybe you only need the else statement below
            current_group_windows = [i for i in range(1, group_count_range)]

            for i in range(1, group_count_range):
                user_train_options.append(['OPTIVagRound-Bold', '##000000', True, default_train_text_size])
            for i in range(1, group_count_range):
                user_name_options.append(['OPTIVagRound-Bold', '#000000', True, default_name_text_size])
            for i in range(1, group_count_range):
                preview_train_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_train_text_size])
            for i in range(1, group_count_range):
                preview_name_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_name_text_size])

            make_dialogs_for_choosing_group_image(group_count_range)
            make_group_options_buttons(group_count_range)
            dpg.add_texture_registry(tag='texture registry', show=False)
        else:
            for button in current_group_options_buttons:
                dpg.delete_item(button)
                dpg.delete_item(button + " spacer")
            for index, window in enumerate(current_group_windows):
                dpg.delete_item(window)

            user_train_options.clear()
            for i in range(1, group_count_range):
                user_train_options.append(['OPTIVagRound-Bold', '#000000', True, default_train_text_size])
            user_name_options.clear()
            for i in range(1, group_count_range):
                user_name_options.append(['OPTIVagRound-Bold', '#000000', True, default_name_text_size])
            preview_train_options.clear()
            for i in range(1, group_count_range):
                preview_train_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_train_text_size])
            preview_name_options.clear()
            for i in range(1, group_count_range):
                preview_name_options.append(['OPTIVagRound-Bold', [0, 0, 0, 255], True, default_name_text_size])

            current_group_windows = [i for i in range(1, group_count_range)]

            for dialog in current_group_image_dialogs:
                dpg.delete_item(dialog)

            dpg.delete_item('create')
            dpg.delete_item('texture registry')
            current_group_image_dialogs.clear()
            current_group_options_buttons.clear()
            make_dialogs_for_choosing_group_image(group_count_range)
            make_group_options_buttons(group_count_range)
            dpg.add_texture_registry(tag='texture registry', show=False)


# Preview functions and customization functions

def callback_for_selected_group_image(sender, app_data):
    file_path = app_data['file_path_name']
    group_count_range = group_count[0] + 1  # group_count range
    for group_number in range(1, group_count_range):
        if sender == f'dialog_for_choosing_image_for_group {str(group_number)}':
            current_images[group_number - 1] = file_path
            add_a_static_texture(group_number, file_path)
            create_background_preview(group_number)

def add_a_static_texture(group_number, file_path):
    try:
        width, height, channels, data = dpg.load_image(file_path)
        dpg.add_static_texture(width, height, data, tag=f'group {group_number} image', parent='texture registry')
        print("Added new texture")
    except SystemError:
        dpg.delete_item(f'group {group_number} draw list')
        dpg.delete_item(f'group {group_number} image')
        width, height, channels, data = dpg.load_image(file_path)
        dpg.add_static_texture(width, height, data, tag=f'group {group_number} image', parent='texture registry')
        print("Deleted and added new texture")

def change_train_text_size(sender, app_data, user_data):
    user_train_options[user_data][3] = int(app_data)
    create_background_preview(user_data + 1)
    
def change_train_text_color(sender, app_data, user_data):
    rgba_value_float = dpg.get_value(f'group {user_data + 1} train color picker')
    rgba_value_int = [round(item) for item in rgba_value_float]
    #preview_train_options[user_data][1] = rgba_value_int

    rgba_value_float.pop()
    rgb_value_int = [round(item) for item in rgba_value_float]
    user_train_options[user_data][1] = rgb_to_hex(rgb_value_int)
    create_background_preview(user_data + 1)

def change_name_text_size(sender, app_data, user_data):
    user_name_options[user_data][3] = int(app_data)
    create_background_preview(user_data + 1)
    
def change_name_text_color(sender, app_data, user_data):
    rgba_value_float = dpg.get_value(f'group {user_data + 1} name color picker')
    rgba_value_int = [round(item) for item in rgba_value_float]
    #preview_name_options[user_data][1] = rgb_value_int_alpha

    rgba_value_float.pop()
    rgb_value_int = [round(item) for item in rgba_value_float]
    user_name_options[user_data][1] = rgb_to_hex(rgb_value_int)
    create_background_preview(user_data + 1)

def train_bold_option(sender, app_data, user_data):
    if app_data is True:
        user_train_options[user_data][2] = True
        preview_train_options[user_data][2] = True
    else:
        user_train_options[user_data][2] = False
        preview_train_options[user_data][2] = False

def name_bold_option(sender, app_data, user_data):
    if app_data is True:
        user_name_options[user_data][2] = True
        preview_name_options[user_data][2] = True
    else:
        user_name_options[user_data][2] = False
        preview_name_options[user_data][2] = False
            
def create_background_preview(group_number):
    name_color = hex_to_rgb(user_name_options[group_number - 1][1])
    name_size = (user_name_options[group_number - 1][3])*2.5
    train_color = [round(color_value) for color_value in hex_to_rgb(user_train_options[group_number - 1][1])]
    train_size = (user_train_options[group_number - 1][3])*2.5
    try:
        dpg.delete_item(f'group {group_number} draw list')
        dpg.add_drawlist(tag=f'group {group_number} draw list', width=450, height=300, parent=f'group {group_number} options window') # 450x300 is a ratio of the Excel dimensions (300x200)
        dpg.draw_image(f'group {group_number} image', (0, 0), (450, 300), parent=f'group {group_number} draw list')
        name_text = dpg.draw_text(((220 - name_size), (180 - name_size)), "Mogan", color=name_color, size=name_size, parent=f'group {group_number} draw list')
        train_text = dpg.draw_text((20, 30), "Train Station", color=train_color, size=train_size, parent=f'group {group_number} draw list')
        dpg.bind_item_font(name_text, name_font)
        dpg.bind_item_font(train_text, name_font)
    except SystemError:
        dpg.delete_item(f'group {group_number} draw list')
        dpg.add_drawlist(tag=f'group {group_number} draw list', width=450, height=300, parent=f'group {group_number} options window')
        name_text = dpg.draw_text(((220 - name_size), (180 - name_size)), "Mogan", color=name_color, size=name_size, parent=f'group {group_number} draw list')
        train_text = dpg.draw_text((20, 30), "Train Station", color=train_color, size=train_size, parent=f'group {group_number} draw list')
        dpg.bind_item_font(name_text, name_font)
        dpg.bind_item_font(train_text, name_font)

def callback_for_selected_back_side_image(sender, app_data):
    global current_back_side_image
    current_back_side_image = app_data['file_path_name']

def callback_for_selected_logo(sender, app_data):
    global current_logo
    current_logo = app_data['file_path_name']

def resolution_picker(sender, app_data):
    print(app_data)


# Dialog functions

def show_dialog(sender, app_data, user_data):
    dpg.show_item(f'dialog_for_choosing_image_for_group {str(user_data)}')

def make_dialogs_for_choosing_group_image(group_count_range):
    for dialog_number in range(1, group_count_range):
        current_group_image_dialogs.append(f'dialog_for_choosing_image_for_group {str(dialog_number)}')
    for group_number in range(1, group_count_range):
        dpg.add_file_dialog(label=f'dialog_for_choosing_image_for_group {str(group_number)}', directory_selector=False, show=False, callback=callback_for_selected_group_image, cancel_callback=cancel_callback, tag=f'dialog_for_choosing_image_for_group {str(group_number)}', width=800, height=400)
        dpg.add_file_extension(".jpg", parent=f'dialog_for_choosing_image_for_group {str(group_number)}', color=(25, 199, 230))
        dpg.add_file_extension(".png", parent=f'dialog_for_choosing_image_for_group {str(group_number)}', color=(25, 199, 230))

def show_group_options_window(sender, app_data, user_data):
    dpg.show_item(f'group {user_data} options window')

def make_group_options_buttons(group_count_range):
    for button_number in range(1, group_count_range):
        current_group_options_buttons.append(f'group {str(button_number)} options')
    for index, button_tag in enumerate(current_group_options_buttons):
        group = index + 1
        dpg.add_button(tag=button_tag, label=button_tag, parent='Main', callback=show_group_options_window, user_data=group, indent=200)
        dpg.add_spacer(tag=f'{button_tag} spacer', height=10, parent='Main')
        current_group_windows[index] = f'group {group} options window'
        dpg.add_window(tag=f'group {group} options window', label=f'group {group} options', show=False, width= 465, height=800, pos=[500,0])

        # Add new group options items here

        dpg.add_spacer(height=10, parent=f'group {group} options window')
        dpg.add_button(indent=80, tag=f'Upload group {group} background image', label=f'Upload group {group} background image', parent=f'group {group} options window', callback=show_dialog, user_data=group)
        dpg.add_spacer(height=10, parent=f'group {group} options window')
        dpg.add_separator(parent=f'group {group} options window')
        dpg.add_spacer(height=10, parent=f'group {group} options window')

        #/ Options set (train options + name options)
        dpg.add_group(tag=f'options set for group {group}', parent=f'group {group} options window', horizontal=True)

        #// train options
        dpg.add_group(tag=f'train options set for group {group}', parent=f'options set for group {group}')

        dpg.add_text(default_value='Train Group', parent=f'train options set for group {group}')
        dpg.add_spacer(height=10, parent=f'train options set for group {group}')
        dpg.add_color_picker(tag=f'group {group} train color picker', parent=f'train options set for group {group}', callback=change_train_text_color, user_data=index, width=130, height=100)
        dpg.add_spacer(height=10, parent=f'train options set for group {group}')
        #dpg.add_checkbox(tag=f'group {group} train bold option', label='Bold text', parent=f'train options set for group {group}', callback=train_bold_option, user_data=index, default_value=True)
        dpg.add_spacer(height=10, parent=f'train options set for group {group}')
        dpg.add_text(default_value='Size:', parent=f'train options set for group {group}')
        dpg.add_radio_button(('14', '18', '20', '24'), parent=f'train options set for group {group}', callback=change_train_text_size, horizontal=True, user_data=index, default_value='20')
        dpg.add_spacer(height=10, parent=f'train options set for group {group}')
        dpg.add_separator(parent=f'train options set for group {group}')
        dpg.add_spacer(height=10, parent=f'train options set for group {group}')

        #// name options
        dpg.add_group(tag=f'name options set for group {group}', parent=f'options set for group {group}')

        dpg.add_text(default_value='Name', parent=f'name options set for group {group}')
        dpg.add_spacer(height=10, parent=f'name options set for group {group}')
        dpg.add_color_picker(tag=f'group {group} name color picker', parent=f'name options set for group {group}', callback=change_name_text_color, user_data=index, width=130, height=100)
        dpg.add_spacer(height=10, parent=f'name options set for group {group}')
        #dpg.add_checkbox(tag=f'group {group} name bold option', label='Bold text', parent=f'name options set for group {group}', callback=name_bold_option, user_data=index, default_value=True)
        dpg.add_spacer(height=10, parent=f'name options set for group {group}')
        dpg.add_text(default_value='Size:', parent=f'name options set for group {group}')
        dpg.add_radio_button(('24', '32', '36', '44'), parent=f'name options set for group {group}', callback=change_name_text_size, horizontal=True, user_data=index, default_value='36')


    dpg.add_button(tag='create', label='Create name tags', parent='Main', callback=create_name_tags, user_data=path_and_sheet_and_count, indent=200)


# Conversion functions

def rgb_to_hex(rgb_tuple):
    return '#' + ''.join(f'{i:02X}' for i in rgb_tuple)

def hex_to_rgb(hex_tuple):
    hex_tuple = hex_tuple.lstrip('#')
    return tuple(int(hex_tuple[i:i+2], 16) for i in (0, 2, 4))



# Not yet finished functions

def cancel_callback(sender, app_data): # What could you use this for?
    print("Sender: ", sender)
    print("App Data: ", app_data)

def auto_resize(sender, app_data, user_data): # Make optional image resize function
    pass



# With area

with dpg.file_dialog(directory_selector=False, show=False, callback=selected_excel_file_callback, cancel_callback=cancel_callback, id="choose_excel_file_dialog", width=800 ,height=600):
    dpg.add_file_extension(".xlsx", color=(25, 199, 230))

with dpg.file_dialog(directory_selector=False, show=False, callback=callback_for_selected_back_side_image, cancel_callback=cancel_callback, id='choose_back_side_image_dialog', width=800, height=600):
    dpg.add_file_extension(".jpg", color=(20, 300, 240))
    dpg.add_file_extension(".png", color=(25, 199, 230))

with dpg.file_dialog(directory_selector=False, show=False, callback=callback_for_selected_logo, cancel_callback=cancel_callback, id='choose_logo_dialog', width=800, height=600):
    dpg.add_file_extension(".jpg", color=(25, 199, 240))
    dpg.add_file_extension(".png", color=(25, 199, 230))

with dpg.font_registry():
    # first argument ids the path to the .ttf or .otf file
    main_font = dpg.add_font('OPTIVagRound-Bold.otf', 22)
    name_font = dpg.add_font('OPTIVagRound-Bold.otf', 100)

with dpg.window(tag='Main'):

    dpg.bind_font(main_font)
    dpg.add_spacer(height=10)
    dpg.add_text('Name Tag Creator', color=[232, 163, 33], indent=200)
    dpg.add_spacer(height=10)

    dpg.add_separator()
    dpg.add_spacer(height=10)

    dpg.add_text('Monitor resolution', color=[0, 90, 200], indent=200)
    dpg.add_radio_button(items=['1080x1920', '1440x1080'], label='Monitor resolution:', tag='resolution picker', callback=resolution_picker, indent=200)
    dpg.add_spacer(height=20)

    dpg.add_text('Activity groups excel file', color=[0, 90, 200], indent=200)
    dpg.add_button(label='Select', callback=lambda: dpg.show_item('choose_excel_file_dialog'), indent=200)
    dpg.add_spacer(height=10)

    #dpg.add_checkbox(label='Image auto resize', tag='Image auto resize', callback=auto_resize)
    #dpg.add_spacer(height=20)

    dpg.add_text('Image for back side of nametags', color=[0, 90, 200], indent=200)
    dpg.add_button(label='Select', callback=lambda: dpg.show_item('choose_back_side_image_dialog'), indent=200)
    dpg.add_spacer(height=10)

    #dpg.add_button(label='Select logo image', callback=lambda: dpg.show_item('choose_logo_dialog'))
    #dpg.add_spacer(height=40)



dpg.setup_dearpygui()
dpg.show_viewport()
dpg.set_primary_window('Main', True)
dpg.start_dearpygui()
dpg.destroy_context()